import streamlit as st
import pandas as pd
import json
import os
import io
from datetime import date, datetime

# ──────────────────────────────────────────────
# CONFIGURACIÓN (Adaptada para Local y Nube)
# ──────────────────────────────────────────────
INVENTARIO_PATH = "inventario.xlsx"
REGISTROS_PATH  = "registros_conteo.json"
EXCEL_REGISTROS = "registros_conteo.xlsx"
CONTEOS = [1, 2, 3, 4, 5]  # Configuración de conteos numerados
UM_OPCIONES = ["UNIDADES", "KILOGRAMOS", "MOLDES", "PLANCHA"]  # Opciones disponibles de Unidad de Medida

st.set_page_config(page_title="Sistema de Dosimetría", page_icon="🧪", layout="wide")

# ──────────────────────────────────────────────
# CARGA Y PERSISTENCIA
# ──────────────────────────────────────────────
@st.cache_data(ttl=60)
def cargar_inventario():
    try:
        df = pd.read_excel(INVENTARIO_PATH)
        df.columns = [c.strip().upper() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"No se pudo cargar el inventario: {e}")
        return pd.DataFrame(columns=["CÓDIGO", "INSUMO", "UM"])

def guardar_inventario(df: pd.DataFrame):
    try:
        df.to_excel(INVENTARIO_PATH, index=False)
        cargar_inventario.clear()
    except PermissionError:
        st.error("⚠️ No se pudo guardar porque **inventario.xlsx** está abierto en Excel. Ciérralo e intenta de nuevo.")
        st.stop()

def cargar_registros() -> dict:
    if os.path.exists(REGISTROS_PATH):
        with open(REGISTROS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_registros(data: dict):
    with open(REGISTROS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    guardar_excel_registros(data)

def guardar_excel_registros(data: dict):
    if not data:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        fechas  = sorted(set(v["fecha"] for v in data.values()), reverse=True)
        color_h = "1A3A5C"
        color_s = "2C6FB5"
        thin    = Side(style="thin", color="CCCCCC")
        borde   = Border(left=thin, right=thin, top=thin, bottom=thin)

        for fecha in fechas:
            regs = sorted([v for v in data.values() if v["fecha"] == fecha], key=lambda x: x["insumo"])
            ws   = wb.create_sheet(title=fecha.replace("-", ""))
            ws.merge_cells("A1:I1")
            ws["A1"] = f"REGISTRO DE CONTEO — {fecha}"
            ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
            ws["A1"].fill      = PatternFill("solid", fgColor=color_h)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28
            
            encabezados = ["CÓDIGO","INSUMO","UM","CONTEO 1","CONTEO 2","CONTEO 3","CONTEO 4","CONTEO 5","TOTAL"]
            for col, h in enumerate(encabezados, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font      = Font(name="Arial", bold=True, color="FFFFFF")
                c.fill      = PatternFill("solid", fgColor=color_s)
                c.alignment = Alignment(horizontal="center")
                c.border    = borde
            ws.row_dimensions[2].height = 20
            for ri, reg in enumerate(regs, 3):
                vals = [reg.get("codigo",""), reg.get("insumo",""), reg.get("um",""),
                        reg.get("mesas",{}).get("1",0), reg.get("mesas",{}).get("2",0),
                        reg.get("mesas",{}).get("3",0), reg.get("mesas",{}).get("4",0),
                        reg.get("mesas",{}).get("5",0), reg.get("total",0)]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.font      = Font(name="Arial", size=10, bold=(ci==9), color=(color_h if ci==9 else "000000"))
                    c.border    = borde
                    c.alignment = Alignment(horizontal="left" if ci==2 else "center")
                if ri % 2 == 0:
                    for ci in range(1,10):
                        ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="EEF2F7")
            uf = 2 + len(regs) + 1
            ws.merge_cells(f"A{uf}:C{uf}")
            ws[f"A{uf}"] = "TOTAL GENERAL"
            ws[f"A{uf}"].font      = Font(name="Arial", bold=True, color="FFFFFF")
            ws[f"A{uf}"].fill      = PatternFill("solid", fgColor=color_h)
            ws[f"A{uf}"].alignment = Alignment(horizontal="center")
            for ci in range(4,10):
                l = get_column_letter(ci)
                c = ws.cell(row=uf, column=ci, value=f"=SUM({l}3:{l}{uf-1})")
                c.font      = Font(name="Arial", bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=color_h)
                c.alignment = Alignment(horizontal="center")
                c.border    = borde
            for ci, w in enumerate([14,40,12,12,12,12,12,12,12], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        wb.save(EXCEL_REGISTROS)
    except Exception:
        pass

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def registros_a_df(data: dict) -> pd.DataFrame:
    if not data:
        return pd.DataFrame()
    rows = []
    for v in data.values():
        row = {"Fecha": v.get("fecha",""), "Código": v.get("codigo",""),
               "Insumo": v.get("insumo",""), "UM": v.get("um","")}
        for c_num in CONTEOS:
            row[f"Conteo {c_num}"] = v.get("mesas",{}).get(str(c_num), 0)
        row["Total"] = v.get("total", 0)
        rows.append(row)
    df = pd.DataFrame(rows)
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    return df

def excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active
    color_s = "2C6FB5"; color_h = "1A3A5C"
    thin = Side(style="thin", color="CCCCCC")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = df.columns.tolist()
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=color_s)
        c.alignment = Alignment(horizontal="center")
        c.border = borde
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = borde
            c.alignment = Alignment(horizontal="left" if ci==3 else "center")
        if ri % 2 == 0:
            for ci in range(1, len(cols)+1):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="EEF2F7")
    for ci in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()

# ──────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #f5f7fa; }
    .block-container { padding-top: 1.5rem; }
    h1 { color: #1a3a5c; }
    h2, h3 { color: #2c5282; }
    .metric-box {
        background: white; border-radius: 10px; padding: 16px 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center;
    }
    .total-box {
        background: linear-gradient(135deg,#1a3a5c,#2c6fb5); color: white;
        border-radius: 12px; padding: 20px; text-align: center;
        font-size: 2rem; font-weight: 700;
        box-shadow: 0 4px 12px rgba(44,111,181,0.3);
    }
    div[data-testid="stTabs"] button { font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# CABECERA
# ──────────────────────────────────────────────
st.title("Sistema de Inventario en Planta")
st.markdown("---")

# ──────────────────────────────────────────────
# PESTAÑAS PRINCIPALES (MODIFICADO: Reportes y Gestión ocultos)
# ──────────────────────────────────────────────
tab1, tab2 = st.tabs([
    "📋 Registro de Conteo",
    "🔍 Consulta por Fecha"
])

# ══════════════════════════════════════════════
# TAB 1 — REGISTRO DE CONTEO
# ══════════════════════════════════════════════
with tab1:
    df_inv    = cargar_inventario()
    registros = cargar_registros()
    st.subheader("📋 Registro de Conteo General")

    col_fecha, col_buscar = st.columns([2, 4])
    with col_fecha:
        fecha_sel = st.date_input("📅 Fecha de conteo", value=date.today(), key="fecha_reg")
        fecha_str = fecha_sel.strftime("%Y-%m-%d")
    with col_buscar:
        insumo_sel = st.selectbox(
            "🔍 Buscar y seleccionar insumo",
            ["-- Seleccione un insumo --"] + df_inv["INSUMO"].tolist(),
            key="insumo_sel", help="Escribe directamente aquí para filtrar"
        )

    if insumo_sel != "-- Seleccione un insumo --":
        fila          = df_inv[df_inv["INSUMO"] == insumo_sel].iloc[0]
        codigo        = fila["CÓDIGO"]
        um_inventario = fila.get("UM", "")

        clave         = f"{fecha_str}__{codigo}"
        existente     = registros.get(clave, {})
        mesas_previas = existente.get("mesas", {str(c_num): 0 for c_num in CONTEOS})

        # Determinar el valor por defecto de la UM: prioriza lo ya guardado para
        # este registro; si no existe, intenta usar la UM del inventario; si
        # ninguna coincide con las opciones disponibles, usa la primera opción.
        um_guardada = existente.get("um", um_inventario)
        if um_guardada in UM_OPCIONES:
            idx_um = UM_OPCIONES.index(um_guardada)
        else:
            idx_um = 0

        c1, c2, c3 = st.columns(3)
        c1.markdown(f"<div class='metric-box'>🔑 <b>Código</b><br>{codigo}</div>", unsafe_allow_html=True)
        c2.markdown(f"<div class='metric-box'>📦 <b>Insumo</b><br>{insumo_sel}</div>", unsafe_allow_html=True)
        with c3:
            um = st.selectbox("⚖️ UM", UM_OPCIONES, index=idx_um, key=f"um_sel_{clave}")
        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("### 🔢 Desglose de Conteos")
        cols_conteos = st.columns(5)
        valores_conteo = {}
        for i, c_num in enumerate(CONTEOS):
            with cols_conteos[i]:
                v = st.number_input(f"Conteo {c_num}", min_value=0.0,
                                    value=float(mesas_previas.get(str(c_num), 0)),
                                    step=0.5, key=f"conteo_{c_num}_{clave}")
                valores_conteo[str(c_num)] = v

        total = sum(valores_conteo.values())
        st.markdown(f"<div class='total-box'>TOTAL DE CONTEO: {total:,.2f} {um}</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        col_guard, col_eliminar = st.columns([3, 1])
        with col_guard:
            if st.button("💾 Guardar Registro", use_container_width=True, type="primary"):
                registros[clave] = {
                    "fecha": fecha_str, "codigo": codigo, "insumo": insumo_sel,
                    "um": um, "mesas": valores_conteo, "total": total,
                    "updated": datetime.now().isoformat()
                }
                guardar_registros(registros)
                st.success(f"✅ Registro guardado para **{insumo_sel}** el **{fecha_str}**")
                st.info("📊 Excel actualizado: `registros_conteo.xlsx`")
                st.balloons()
        with col_eliminar:
            if clave in registros:
                if st.button("🗑️ Eliminar Registro", use_container_width=True, type="secondary"):
                    del registros[clave]
                    guardar_registros(registros)
                    st.warning(f"⚠️ Registro eliminado para **{insumo_sel}** el **{fecha_str}**")
                    st.rerun()

    st.markdown("---")
    registros_dia = [v for v in registros.values() if v.get("fecha") == fecha_str]
    if registros_dia:
        st.markdown(f"### 📑 Registros del día: {fecha_str} ({len(registros_dia)} insumos)")
        df_dia = pd.DataFrame(registros_dia)[["codigo","insumo","mesas","total","um"]]
        for c_num in CONTEOS:
            df_dia[f"Conteo {c_num}"] = df_dia["mesas"].apply(lambda x: x.get(str(c_num), 0))
        df_dia = df_dia.drop(columns=["mesas"]).rename(columns={
            "codigo":"Código","insumo":"Insumo","total":"Total","um":"UM"})
        
        columnas_ordenadas = ["Código", "Insumo", "UM"] + [f"Conteo {c_num}" for c_num in CONTEOS] + ["Total"]
        st.dataframe(df_dia[columnas_ordenadas], use_container_width=True, hide_index=True)
    else:
        st.info(f"📭 No hay registros para el {fecha_str}")

# ══════════════════════════════════════════════
# TAB 2 — CONSULTA POR FECHA
# ══════════════════════════════════════════════
with tab2:
    st.subheader("🔍 Consulta de Inventario por Fecha")
    registros = cargar_registros()
    fechas_disponibles = sorted(set(v["fecha"] for v in registros.values()), reverse=True)

    if not fechas_disponibles:
        st.info("📭 Aún no hay registros guardados.")
    else:
        fecha_consulta     = st.date_input("📅 Seleccione la fecha", value=date.today(), key="fecha_consulta")
        fecha_consulta_str = fecha_consulta.strftime("%Y-%m-%d")
        resultados = [v for v in registros.values() if v.get("fecha") == fecha_consulta_str]

        if resultados:
            st.success(f"✅ **{len(resultados)}** registros para el **{fecha_consulta_str}**")
            df_res = pd.DataFrame(resultados)
            for c_num in CONTEOS:
                df_res[f"Conteo {c_num}"] = df_res["mesas"].apply(lambda x: x.get(str(c_num), 0))
            df_res = df_res.drop(columns=["mesas","updated"], errors="ignore").rename(columns={
                "fecha":"Fecha","codigo":"Código","insumo":"Insumo","total":"Total","um":"UM"})
            df_res = df_res[["Código","Insumo","UM"]+[f"Conteo {c_num}" for c_num in CONTEOS]+["Total"]].sort_values("Insumo")
            st.dataframe(df_res, use_container_width=True, hide_index=True)
            c1, c2 = st.columns(2)
            c1.metric("📦 Insumos contados", len(resultados))
            c2.metric("⚖️ Total general", f"{df_res['Total'].sum():,.2f}")
            st.download_button("📥 Descargar Excel", data=excel_bytes(df_res),
                               file_name=f"conteo_{fecha_consulta_str}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning(f"📭 No hay registros para **{fecha_consulta_str}**")
